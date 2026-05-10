from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.report_service import ReportFilters

FONT_DIR = Path(__file__).resolve().parent.parent / "assets" / "fonts"
FONT_REG_PATH = FONT_DIR / "NotoSansBengali-Regular.ttf"
FONT_BOLD_PATH = FONT_DIR / "NotoSansBengali-Bold.ttf"


def _register_unicode_fonts() -> tuple[str, str]:
    """
    Register Noto Sans Bengali for Latin + Bengali (Bangla) text in PDFs.
    Falls back to Helvetica if font files are missing (Bangla may not render).
    """
    reg_name = "NotoBengali"
    bold_name = "NotoBengali-Bold"
    if not FONT_REG_PATH.is_file():
        return "Helvetica", "Helvetica-Bold"
    try:
        pdfmetrics.getFont(reg_name)
    except KeyError:
        pdfmetrics.registerFont(TTFont(reg_name, str(FONT_REG_PATH)))
    if FONT_BOLD_PATH.is_file():
        try:
            pdfmetrics.getFont(bold_name)
        except KeyError:
            pdfmetrics.registerFont(TTFont(bold_name, str(FONT_BOLD_PATH)))
    else:
        bold_name = reg_name
    return reg_name, bold_name


def _cell_paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    raw = str(text or "")
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    safe = escape(raw).replace("\n", "<br/>")
    return Paragraph(safe, style)


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return dt.strftime("%Y-%m-%d %H:%M")


def format_letters_filter_caption(
    filters: ReportFilters,
    *,
    department_name: str | None = None,
) -> str:
    """Human-readable summary of report filters (matches export query params)."""
    parts: list[str] = []
    if filters.date_from is not None and filters.date_to is not None:
        parts.append(f"Created {filters.date_from} to {filters.date_to}")
    elif filters.date_from is not None:
        parts.append(f"Created from {filters.date_from}")
    elif filters.date_to is not None:
        parts.append(f"Created until {filters.date_to}")
    else:
        parts.append("Dates: all (no range filter)")

    if filters.status is not None:
        parts.append(f"Status: {filters.status.value}")
    else:
        parts.append("Status: any")

    if filters.department_id is not None:
        label = department_name or f"ID {filters.department_id}"
        parts.append(f"Department: {label}")
    else:
        parts.append("Department: all in scope")

    if filters.q:
        parts.append(f"Search: {filters.q}")
    else:
        parts.append("Search: none")

    if filters.from_office:
        parts.append(f"From office: {filters.from_office}")
    else:
        parts.append("From office: none")

    return "  |  ".join(parts)


def build_letters_pdf(
    rows: list[dict[str, str]],
    *,
    title: str = "Letters export (according to filter)",
    filter_caption: str = "",
) -> bytes:
    buf = BytesIO()
    # A4 landscape — fits eight columns like the spreadsheet export.
    page = landscape(A4)
    page_w, _ = page
    lm = 0.5 * inch
    rm = 0.5 * inch
    tm = 0.5 * inch
    bm = 0.5 * inch
    usable_w = page_w - lm - rm

    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=lm,
        rightMargin=rm,
        topMargin=tm,
        bottomMargin=bm,
        title=title[:120],
    )
    font_reg, font_bold = _register_unicode_fonts()

    style_title = ParagraphStyle(
        name="LettersExportTitle",
        fontName=font_bold,
        fontSize=13,
        leading=16,
        alignment=TA_LEFT,
        textColor=colors.black,
        spaceAfter=4,
    )
    style_caption = ParagraphStyle(
        name="LettersExportCaption",
        fontName=font_reg,
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#374151"),
        spaceAfter=8,
    )
    # Spreadsheet-style header: bold black on light grey (see Excel export).
    style_hdr = ParagraphStyle(
        name="LettersExportHdr",
        fontName=font_bold,
        fontSize=9,
        leading=11,
        alignment=TA_LEFT,
        textColor=colors.black,
    )
    style_cell = ParagraphStyle(
        name="LettersExportCell",
        fontName=font_reg,
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )

    story: list = [_cell_paragraph(title, style_title)]
    if filter_caption.strip():
        story.append(_cell_paragraph(filter_caption.strip(), style_caption))
    story.append(Spacer(1, 2))

    header = [
        "Serial",
        "Memo No",
        "Subject",
        "Status",
        "Priority",
        "Department",
        "Received from",
        "Created",
    ]

    # Column widths (points) for A4 landscape; Subject gets the remainder.
    w_serial = 68
    w_memo = 74
    w_status = 72
    w_priority = 52
    w_dept = 76
    w_recv = 86
    w_created = 92
    fixed = w_serial + w_memo + w_status + w_priority + w_dept + w_recv + w_created
    w_subject = max(140.0, float(usable_w) - float(fixed))
    col_widths = [w_serial, w_memo, w_subject, w_status, w_priority, w_dept, w_recv, w_created]

    header_cells = [_cell_paragraph(h, style_hdr) for h in header]
    table_data: list[list[Paragraph]] = [header_cells]
    for r in rows:
        table_data.append(
            [
                _cell_paragraph(r.get("serial_no", ""), style_cell),
                _cell_paragraph(r.get("memo_no", ""), style_cell),
                _cell_paragraph(r.get("subject", ""), style_cell),
                _cell_paragraph(r.get("status", ""), style_cell),
                _cell_paragraph(r.get("priority", ""), style_cell),
                _cell_paragraph(r.get("department", ""), style_cell),
                _cell_paragraph(r.get("received_from", ""), style_cell),
                _cell_paragraph(r.get("created_at", ""), style_cell),
            ]
        )

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                # Same look as spreadsheet: light grey header, bold black via ParagraphStyle.
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()


def build_letters_xlsx(
    rows: list[dict[str, str]],
    *,
    title: str = "Letters export (according to filter)",
    filter_caption: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Letters"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    if filter_caption.strip():
        ws.append([filter_caption.strip()])
        ws["A2"].font = Font(size=10)
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.append([])
    ws.append(
        [
            "Serial",
            "Memo No",
            "Subject",
            "Status",
            "Priority",
            "Department",
            "Received from",
            "Created",
        ]
    )
    header_row = ws.max_row
    header_fill = PatternFill(start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid")
    header_font = Font(bold=True, color="FF000000", size=11)
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(
            [
                r.get("serial_no", ""),
                r.get("memo_no", ""),
                r.get("subject", ""),
                r.get("status", ""),
                r.get("priority", ""),
                r.get("department", ""),
                r.get("received_from", ""),
                r.get("created_at", ""),
            ]
        )
    zebra_fill = PatternFill(start_color="FFF9FAFB", end_color="FFF9FAFB", fill_type="solid")
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row)):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def letter_row_from_model(letter) -> dict[str, str]:
    dept_name = ""
    if letter.department is not None:
        dept_name = getattr(letter.department, "name", "") or ""
    memo = getattr(letter, "memo_no", None) or ""
    return {
        "serial_no": letter.serial_no,
        "memo_no": memo,
        "subject": letter.subject or "",
        "status": letter.status.value if hasattr(letter.status, "value") else str(letter.status),
        "priority": letter.priority.value if hasattr(letter.priority, "value") else str(letter.priority),
        "department": dept_name,
        "received_from": letter.received_from or "",
        "created_at": _fmt_dt(letter.created_at),
    }
